import openpyxl
import itertools

from numpy import array
import numpy as np
#=====================================================================
# only works for multi dimensional rows Rows
class cax_xlsx:
    def LoadRange2D(sRangeName):  ## min size 2D
        #nRangeRow = 0,1,2...
         #[('Sheet1', '$B$7:$E$10')]
        #removing the $ from the address
        print('====< Load Range: LoadRange(sRangeName) >=========================================')
        ##------------------------------------------------------
        arrayRange=[]
        address = list(wb.defined_names[sRangeName].destinations)
        for sheetname, cellAddress in address:
            cellAddress = cellAddress.replace('$','')
            #'B7:E10'
            print('Range Name= ',sRangeName,'; cellAddress= ',cellAddress)
        ##............................................................................................
        worksheet = wb[sheetname]
        a = array(worksheet[cellAddress])
        nSize=len(a.shape)
        print('cellAddress:',cellAddress,'  Size of Range: ', nSize , '   Shape:  ' , a.shape)
        print('Rows:',a.shape[0],' Columns:',a.shape[1])

        print('----< Create array >-------------------------------------')
        nRows=a.shape[0]
        nColumns=a.shape[1]
        lRows=list(range(nRows))
        lColumns=list(range(nColumns))

        ltCells=list(itertools.product(lRows,lColumns))   #2D
        print('----2D list of cell tuples---------')
        print(ltCells)

        arrayRange =[[0]*nColumns]*nRows
        arrayRange = np.array(arrayRange)


        print('----of all cells---------')
        for item in ltCells:
            nRow=item[0]
            nColumn=item[1]
            rCell=worksheet[cellAddress][nRow][nColumn].value

            arrayRange[nRow][nColumn]=rCell

            print(item, 'Row=', nRow,' Column=', nColumn, rCell)

        print('----< /Create array >-------------------------------------')

        print('----2D list of Lists---------')
        print(arrayRange)

        # using list comprehension
        # convert list of tuples to list of list
##        lCells = [list(ele) for ele in ltCells]

        print('====</ Load Range >============================')

##        arrayRange =[[0]*nColumns]*nRows


        return arrayRange

    #========================================================
    def PopulateRange2D(arrayTable,sRangeName,Overflow):  ## min size 2D


## to do
## populate dimensions determined by arrytable not destination range
##
##
        #nRangeRow = 0,1,2...
         #[('Sheet1', '$B$7:$E$10')]
        #removing the $ from the address
        print('====< Load Range: LoadRange(sRangeName) >=========================================')
        ##------------------------------------------------------
        arrayRange=[]
        address = list(wb.defined_names[sRangeName].destinations)
        for sheetname, cellAddress in address:
            cellAddress = cellAddress.replace('$','')
            #'B7:E10'
            print('Range Name= ',sRangeName,'; cellAddress= ',cellAddress)
        ##............................................................................................
        worksheet = wb[sheetname]
        a = array(worksheet[cellAddress])
        nSize=len(a.shape)
        print('cellAddress:',cellAddress,'  Size of Range: ', nSize , '   Shape:  ' , a.shape)
        print('Rows:',a.shape[0],' Columns:',a.shape[1])

        print('----< Create array >-------------------------------------')
        nRows=a.shape[0]
        nColumns=a.shape[1]
        lRows=list(range(nRows))
        lColumns=list(range(nColumns))

        ltCells=list(itertools.product(lRows,lColumns))   #2D
        print('----2D list of cell tuples---------')
        print(ltCells)

        arrayRange =[[0]*nColumns]*nRows
        arrayRange = np.array(arrayRange)


        print('----of all cells---------')
        for item in ltCells:
            nRow=item[0]
            nColumn=item[1]
            rCell=arrayTable[nRow][nColumn]

            if rCell != None:
             worksheet[cellAddress][nRow][nColumn].value=rCell

            print(item, 'Row=', nRow,' Column=', nColumn, rCell)

        print('----< /Create array >-------------------------------------')

        print('----2D list of Lists---------')
        print(arrayRange)

        # using list comprehension
        # convert list of tuples to list of list
##        lCells = [list(ele) for ele in ltCells]

        print('====</ Load Range >============================')

##        arrayRange =[[0]*nColumns]*nRows


        return arrayRange

    #========================================================
    def FillRange(nRangeRowCol,sRangeName,oWorkBook,lList):
        #nRangeRow = 0,1,2...
        address = list(oWorkBook.defined_names[sRangeName].destinations)
        #[('Sheet1', '$B$7:$E$10')]
        #removing the $ from the address
        #------------------------------------------------------
        for sheetname, cellAddress in address:
            cellAddress = cellAddress.replace('$','')
            #'B7:E10'

        worksheet = oWorkBook[sheetname]

##        print('-----')
##        print(worksheet[cellAddress])
##        print('.......................................')
##        print(worksheet[cellAddress][1])
##        print('.......................................')
##        print(worksheet[cellAddress][1][0])
##        print('.......................................')

        # determine verical array or horrizontal
        a = array(worksheet[cellAddress])
        nSize=len(a.shape)

        if nSize>0 :
            tShape=a.shape

        print('cellAddress:',cellAddress,'  Size of Range: ', nSize , '   Shape:  ' , a.shape)

        ##............................................................................................
        ##...THEORY  .................................................................................
        ## cellAddress: 02  Size of Range:  0    Shape:   (2, 10)
        ## cellAddress: O5:X6   Size of Range:  2    Shape:   (2, 10)
        ## cellAddress: E17:E29   Size of Range:  2    Shape:   (13, 1)
        ##............................................................................................

        if nSize>0 :
            nRows=tShape[0]
            nCols=tShape[1]
        ##............................................................................................
            if tShape[0] ==2 or tShape[0] ==1 :
                nRangeXdim=tShape[1] -1           #len(worksheet[cellAddress][1])-1
                i=0   # NOT i=0  because LEVEL Heading introduced
                for item in lList:
                        i=i+1
                        if i < nRangeXdim :
                            worksheet[cellAddress][nRangeRowCol][i].value=item
                            print('Item= ',item)
                        else:
                            worksheet[cellAddress][nRangeRowCol][nRangeXdim].value='DATA OVERFLOW'
        ##............................................................................................

        return oWorkBook
    #========================================================
    def OpenWorkbook(sPathNameWorkBook):
        print('====OpenWorkbook= ',sPathNameWorkBook)
        wb = openpyxl.load_workbook(sPathNameWorkBook,data_only=True) # otherwise formula is displayed
        return wb
    #========================================================
    def SaveWorkbook(sPathNameWorkBook,wb):
        print('====SaveWorkbook')
        wb.save(sPathNameWorkBook)
    #========================================================

#========================================================
#---------------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------------
# Testing!
#---------------------------------------------------------------------------------------------------

if __name__ == '__main__':

    #----OPEN WORKBOOK---------------------------------------------------
    print('====MAIN=========================================')


    sWorkBookInput='TEST_CAX_PY.xlsm'
    sWorkBookOutput='TEST_CAX_PY_out.xlsm'

    sCellName1='RNG_X'

    sRangeName1='RNG_Matrix'  # 2D
    sRangeName2='RNG_Matrix_B'  # 2D

    sRangeName3='rng_FixedSize'
    sRangeName4='rng_FixedSize'

    print('====OPEN WORKBOOK XLSX')
    wb = cax_xlsx.OpenWorkbook(sWorkBookInput)

    rngRange01=cax_xlsx.LoadRange2D(sRangeName1)

    rngRange01=cax_xlsx.PopulateRange2D(rngRange01,sRangeName2,True)



##    rngRange01=cax_xlsx.LoadRange2D(sCellName1)      #.......single cell

    print('==============================================')
    print('The Range:')
    for items in rngRange01:
        print(items)

    print('==============================================')
    print('====SAVE WORKBOOK XLSX')
    cax_xlsx.SaveWorkbook(sWorkBookOutput,wb)

    print('==============================================')
    print('====EXECUTION COMPLETE===')
    print('==============================================')


